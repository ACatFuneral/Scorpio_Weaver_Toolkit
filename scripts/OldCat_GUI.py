# -*- coding: utf-8 -*-
import os
import re
import sys
import subprocess
import json
import threading
import shutil
import time
import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext, ttk
from collections import Counter
from openpyxl import Workbook, load_workbook
import pandas as pd

# ==================== 界面逻辑 ====================

class OldCatApp:
    def __init__(self, root):
        self.root = root
        self.root.title("哈基猫 Ren'Py 盒子 v8.2")
        self.root.geometry("750x700")

        # style setup
        style = ttk.Style()
        style.configure("TNotebook.Tab", font=("Arial", 11, "bold"), padding=[10, 5])

        # --- A. 顶部全局设置 (共享) ---
        frame_global = tk.LabelFrame(root, text="第一步：全局项目配置 (对所有功能生效)", padx=10, pady=10, bg="#f0f0f0")
        frame_global.pack(fill="x", padx=10, pady=5)

        # 游戏目录
        tk.Label(frame_global, text="游戏根目录:", bg="#f0f0f0").grid(row=0, column=0, sticky="e")
        self.entry_path = tk.Entry(frame_global, width=55)
        self.entry_path.grid(row=0, column=1, padx=5, pady=2)
        tk.Button(frame_global, text="浏览...", command=self.browse_folder).grid(row=0, column=2, padx=2)

        # 语言
        tk.Label(frame_global, text="目标语言 (tl目录名):", bg="#f0f0f0").grid(row=1, column=0, sticky="e")
        self.entry_lang = tk.Entry(frame_global, width=20)
        self.entry_lang.insert(0, "Chinese")
        self.entry_lang.grid(row=1, column=1, sticky="w", padx=5, pady=2)

        # --- B. 功能分页区 ---
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill="both", expand=True, padx=10, pady=5)

        # 分页1：文本提取
        self.tab_extract = tk.Frame(self.notebook)
        self.notebook.add(self.tab_extract, text=" 📂 功能1：文本提取 ")
        self.setup_extract_tab()

        # 分页2：Tag 保护
        self.tab_protect = tk.Frame(self.notebook)
        self.notebook.add(self.tab_protect, text=" 🛡️ 功能2：Tag 保护 (禁翻) ")
        self.setup_protect_tab()

        # --- C. 底部日志 ---
        frame_log = tk.LabelFrame(root, text="运行日志", padx=5, pady=5)
        frame_log.pack(fill="both", expand=True, padx=10, pady=5)
        self.text_log = scrolledtext.ScrolledText(frame_log, height=12)
        self.text_log.pack(fill="both", expand=True)
        
        # 重定向输出
        sys.stdout = TextRedirector(self.text_log)
        
        # 自动检查依赖
        self.check_deps()

    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.entry_path.delete(0, tk.END)
            self.entry_path.insert(0, folder)

    def check_deps(self):
        try:
            import pandas; import openpyxl; import yaml
            print("✔ 环境检查通过，全系统就绪。")
        except ImportError: print("⚠️ 正在准备环境...")

    # ========== Tab 1: 提取功能界面 ==========
    def setup_extract_tab(self):
        frame = tk.LabelFrame(self.tab_extract, text="提取设置", padx=20, pady=20)
        frame.pack(fill="both", expand=True, padx=10, pady=10)

        # 模式
        tk.Label(frame, text="提取模式:").grid(row=0, column=0, sticky="e", pady=10)
        self.mode_var = tk.StringVar(value="1")
        modes = [("标准模式 (稳)", "1"), ("标准 + 外部文件", "2"), ("疯狗模式 (全量提取/慎用)", "3")]
        f_radio = tk.Frame(frame)
        f_radio.grid(row=0, column=1, sticky="w")
        for txt, val in modes:
            tk.Radiobutton(f_radio, text=txt, variable=self.mode_var, value=val).pack(side="left", padx=5)

        # 开关
        self.var_emoji = tk.BooleanVar(value=True) # 默认开启
        self.var_ai = tk.BooleanVar(value=True)
        tk.Checkbutton(frame, text="生成 Tag保护/禁翻列表 (推荐开启)", variable=self.var_emoji, fg="blue").grid(row=1, column=1, sticky="w", padx=5)
        tk.Checkbutton(frame, text="生成 AI 提示词文件 (Prompt)", variable=self.var_ai).grid(row=2, column=1, sticky="w", padx=5)

        # 按钮
        self.btn_run_extract = tk.Button(frame, text="🚀 开始提取文本", bg="#4CAF50", fg="white", font=("Arial", 12, "bold"), height=2, command=self.start_extract_thread)
        self.btn_run_extract.grid(row=3, column=0, columnspan=2, sticky="ew", pady=20, padx=50)

    # ========== Tab 2: 保护功能界面 ==========
    def setup_protect_tab(self):
        frame = tk.Frame(self.tab_protect, padx=20, pady=20)
        frame.pack(fill="both", expand=True)

        tk.Label(frame, text="本功能用于防止翻译时损坏代码标签 (如 {w=0.5}, [shake])", fg="gray").pack(pady=5)
        tk.Label(frame, text="原理：代码 -> Emoji (翻译) -> 代码", fg="gray").pack(pady=5)

        # 按钮容器
        btn_frame = tk.Frame(frame)
        btn_frame.pack(pady=20)

        self.btn_encrypt = tk.Button(btn_frame, text="🔒 1. 译前加密\n(将代码替换为 Emoji)", bg="#FF9800", fg="white", font=("Arial", 11, "bold"), width=25, height=3, command=lambda: self.start_protect_thread(1))
        self.btn_encrypt.pack(side="left", padx=20)

        self.btn_decrypt = tk.Button(btn_frame, text="🔓 2. 译后还原\n(将 Emoji 还原为代码)", bg="#2196F3", fg="white", font=("Arial", 11, "bold"), width=25, height=3, command=lambda: self.start_protect_thread(2))
        self.btn_decrypt.pack(side="left", padx=20)

        tk.Label(frame, text="* 注意：操作前会自动备份 tl 文件夹，请放心使用。", fg="red").pack(pady=10)

    # ========== 线程启动器 ==========
    def start_extract_thread(self):
        threading.Thread(target=self.run_extract, daemon=True).start()

    def start_protect_thread(self, mode):
        threading.Thread(target=self.run_protect, args=(mode,), daemon=True).start()

    # ========== 逻辑核心：提取 ==========
    def run_extract(self):
        self.btn_run_extract.config(state="disabled", text="正在提取...")
        game_path = self.entry_path.get().strip()
        lang = self.entry_lang.get().strip()
        
        if not self.validate_path(game_path):
            self.btn_run_extract.config(state="normal", text="🚀 开始提取文本")
            return

        try:
            check_and_install_dependencies()
            target_game_dir = self.fix_game_dir(game_path)
            
            print(f"\n>>> [提取模式] 目标: {target_game_dir}")
            stats = core_logic_extract(target_game_dir, lang, self.mode_var.get(), self.var_emoji.get(), self.var_ai.get())
            
            summary = f"✅ 提取完成！\n\n• 扫描文件：{stats['files']} 个\n• 提取总数：{stats['total']} 条\n  - 人名：{stats['names']}\n  - 对话：{stats['others']}\n  - 替换：{stats['replace']}"
            messagebox.showinfo("哈基猫 Ren'Py 盒子", summary)
        except Exception as e:
            import traceback; print(traceback.format_exc())
            messagebox.showerror("错误", str(e))
        finally:
            self.btn_run_extract.config(state="normal", text="🚀 开始提取文本")

    # ========== 逻辑核心：保护 ==========
    def run_protect(self, mode):
        action_name = "加密" if mode == 1 else "还原"
        self.btn_encrypt.config(state="disabled")
        self.btn_decrypt.config(state="disabled")

        game_path = self.entry_path.get().strip()
        lang = self.entry_lang.get().strip()

        if not self.validate_path(game_path):
            self.reset_protect_btns()
            return

        try:
            # 1. 确定路径
            fixed_game_dir = self.fix_game_dir(game_path) # .../game
            tl_dir = os.path.join(fixed_game_dir, 'tl', lang) # .../game/tl/Chinese
            
            base_dir = os.getcwd()
            # 自动寻找 Excel
            excel_name = "Tag_Protection_Pre(译前).xlsx" if mode == 1 else "Tag_Protection_Post(译后).xlsx"
            excel_path = os.path.join(base_dir, 'translate_output', '3_Emoji_Tools', excel_name)

            if not os.path.exists(tl_dir):
                raise FileNotFoundError(f"找不到翻译目录:\n{tl_dir}\n请检查语言设置是否正确！")
            
            if not os.path.exists(excel_path):
                raise FileNotFoundError(f"找不到映射表:\n{excel_path}\n请先在[功能1]中勾选生成Tag表并运行提取！")

            print(f"\n>>> [{action_name}模式] 启动...")
            print(f"  - 目标目录: {tl_dir}")
            print(f"  - 映射表: {excel_name}")

            # 2. 备份
            print("  >>> 正在创建备份...")
            backup_path = tl_dir + f"_backup_{int(time.time())}"
            shutil.copytree(tl_dir, backup_path)
            print(f"  ✔ 备份已建立: {os.path.basename(backup_path)}")

            # 3. 读取 Excel
            df = pd.read_excel(excel_path, header=None)
            mapping = {}
            # 智能判断标题行
            start_idx = 0
            if "原文" in str(df.iloc[0,0]) or "Tag" in str(df.iloc[0,0]): start_idx = 1
            
            for i in range(start_idx, len(df)):
                k, v = str(df.iloc[i,0]), str(df.iloc[i,1])
                if k != 'nan' and v != 'nan': mapping[k] = v
            
            # 按长度排序
            sorted_map = dict(sorted(mapping.items(), key=lambda item: len(item[0]), reverse=True))

            # 4. 执行替换
            count = 0
            for root, dirs, files in os.walk(tl_dir):
                for file in files:
                    if file.endswith('.rpy'):
                        fpath = os.path.join(root, file)
                        with open(fpath, 'r', encoding='utf-8') as f: content = f.read()
                        new_content = content
                        for k, v in sorted_map.items():
                            if k in new_content: new_content = new_content.replace(k, v)
                        
                        if new_content != content:
                            with open(fpath, 'w', encoding='utf-8') as f: f.write(new_content)
                            print(f"    - 处理: {file}")
                            count += 1
            
            print(f"\n✅ {action_name}完成！共修改 {count} 个文件。")
            messagebox.showinfo("成功", f"{action_name}操作完成！\n修改文件数: {count}\n\n备份路径: {backup_path}")

        except Exception as e:
            import traceback; print(traceback.format_exc())
            messagebox.showerror("保护功能出错", str(e))
        finally:
            self.reset_protect_btns()

    def reset_protect_btns(self):
        self.btn_encrypt.config(state="normal")
        self.btn_decrypt.config(state="normal")

    def validate_path(self, path):
        if not path or not os.path.exists(path):
            messagebox.showerror("路径错误", "请先选择有效的游戏目录！")
            return False
        return True

    def fix_game_dir(self, path):
        # 智能修正，如果选的是根目录，自动加 /game
        if os.path.basename(path) != 'game' and os.path.exists(os.path.join(path, 'game')):
            return os.path.join(path, 'game')
        return path

class TextRedirector:
    def __init__(self, widget):
        self.widget = widget
    def write(self, str):
        self.widget.insert(tk.END, str)
        self.widget.see(tk.END)
    def flush(self): pass

# ==================== 核心逻辑: 提取 (Extract) ====================

def check_and_install_dependencies():
    packages = ["pandas", "openpyxl", "pyyaml"]
    missing = []
    try: import pandas; import openpyxl
    except ImportError: missing.extend(["pandas", "openpyxl"])
    try: import yaml
    except ImportError: 
        if "pyyaml" not in missing: missing.append("pyyaml")
    if not missing: return
    print(f"正在安装依赖: {', '.join(missing)}...")
    subprocess.check_call([sys.executable, "-m", "pip", "install"] + missing)
    global yaml
    import yaml

GRS_EMBEDDED_EMOJIS = ["🌳", "🌴", "🌵", "🌷", "🌹", "🌺", "🌻", "🌼", "🌾", "🌿", "🍀", "🍁", "🍂", "🍃", "🍇", "🍉", "🍊", "🍋", "🍌", "🍎", "🍔", "🍕", "🍟", "🍦", "🍰", "🍷", "🍹", "🍺", "🎀", "🎁", "🎂", "🎈", "🎉", "🎊", "🎓", "🎙️", "🎤", "🎧", "🎨", "🎩", "🎭", "🎮", "🎯", "🎰", "🎱", "🎲", "🎳", "🎵", "🎶", "🎸", "🎹", "🎺", "🎻", "🎼", "🎽", "🎾", "🎿", "🏀", "🏁", "🏂", "🏃", "🏄", "🏅", "🏆", "🏇", "🏈", "🏉", "🏊", "🏋️", "🏐", "🏠", "🏡", "🏢", "🏥", "🏦", "🏪", "🏫", "🏬", "🏰"]

def unescape_string(s):
    if not isinstance(s, str): return str(s)
    return s.replace('\\"', '"').replace("\\'", "'").replace("\\\\", "\\")

def has_chinese(s):
    return bool(re.search(r'[\u4e00-\u9fa5]', s))

def filter_extracted_strings(strings, strict_mode=False):
    filtered_list = []
    file_extensions = ('.mp3', '.png', '.jpg', '.jpeg', '.ogg', '.wav', '.webp', '.gif', '.avi', '.mp4', '.mov', '.webm', '.flv', '.wmv', '.rpy', '.py', '.json', '.yaml', '.yml', '.ttf', '.otf', '.xml', '.csv')
    code_keywords = {'true', 'false', 'none', 'null', 'return', 'jump', 'call', 'label', 'screen', 'style', 'transform', 'image', 'define', 'default', 'init', 'python', 'if', 'else', 'elif', 'for', 'while', 'in', 'and', 'or', 'not', 'pass', 'break', 'continue', 'set', 'get', 'music', 'sound', 'play', 'stop', 'scene', 'show', 'hide', 'with', 'at', 'persistent', 'xalign', 'yalign', 'xpos', 'ypos', 'zoom', 'alpha', 'linear', 'ease'}

    for string in strings:
        if not string or not isinstance(string, str) or string.strip() == '': continue
        if has_chinese(string): continue 
        lower_string = string.lower()
        if any(ext in lower_string for ext in file_extensions) and ' ' not in string: continue
        if string.isdigit(): continue
        if lower_string in code_keywords: continue
        temp_string = re.sub(r'\{.*?\}', '', string)
        temp_string = re.sub(r'\[.*?\]', '', temp_string)
        if not re.search(r'[\u4e00-\u9fa5a-zA-Z0-9]', temp_string) and len(temp_string) > 0: continue
        if (string.startswith('[') and string.endswith(']')) or (string.startswith('{') and string.endswith('}')):
            if not ('{/' in string or ' ' in string): continue
        if string.startswith('#'): continue
        if strict_mode:
            if ' ' not in string and len(string) < 4: continue
            if string.islower() and '_' in string and ' ' not in string: continue
            if string.isupper() and ' ' not in string: continue
            if '/' in string and ' ' not in string: continue
        filtered_list.append(string)
    return filtered_list

def extract_strings_from_rpy(file_path):
    name_strings, text_strings, variable_strings, replace_strings = [], [], [], []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f: content = f.read()
        char_patterns = [r'Character\s*\(\s*(["\'])((?:\\\1|.)*?)\1', r'define\s+\w+\s*=\s*Character\s*\(\s*(["\'])((?:\\\1|.)*?)\1']
        for pattern in char_patterns:
            for match in re.finditer(pattern, content, re.IGNORECASE):
                s = unescape_string(match.group(2))
                if not re.search(r'_\s*\(\s*["\']' + re.escape(s), content): name_strings.append(s)
        text_patterns = [r'\btext\s+(["\'])((?:\\\1|.)*?)\1\s*:', r'\b(text|textbutton|show\s+text)\s+(["\'])((?:\\\2|.)*?)\2', r'renpy\.input\s*\(\s*(["\'])((?:\\\1|.)*?)\1']
        for pattern in text_patterns:
            for match in re.finditer(pattern, content, re.IGNORECASE | re.MULTILINE):
                idx = 3 if pattern == text_patterns[1] else 2
                s = unescape_string(match.group(idx))
                start_pos = match.start()
                line_start = content.rfind('\n', 0, start_pos)
                if not re.search(r'_\s*\(\s*$', content[line_start:start_pos].strip()): text_strings.append(s)
        variable_keywords = [r'default\s+\w+\s*=\s*', r'define\s+\w+\s*=\s*', r'\$\s*\w+\s*=\s*']
        for line in content.split('\n'):
            for keyword in variable_keywords:
                if re.search(keyword, line) and "Character" not in line and not re.search(r'_\s*\(', line):
                    for match in re.finditer(r'(["\'])((?:\\\1|.)*?)\1', line): variable_strings.append(unescape_string(match.group(2)))
            if ('f"' in line or "f'" in line) and not re.search(r'_\s*\(\s*f', line):
                for match in re.finditer(r'f(["\'])((?:\\\1|.)*?)\1', line): replace_strings.append(unescape_string(match.group(2)))
            if re.search(r'^\s*\$\s*(renpy\.notify|csay)\s*\(', line):
                for match in re.finditer(r'(["\'])((?:\\\1|.)*?)\1', line):
                    if not re.search(r'_\s*\(\s*$', line[:match.start()].rstrip()): text_strings.append(unescape_string(match.group(2)))
        for match in re.finditer(r'\btooltip\s*\(\s*(["\'])((?:\\\1|.)*?)\1', content, re.IGNORECASE):
            replace_strings.append(unescape_string(match.group(2)))
    except Exception as e: pass
    return name_strings, text_strings, variable_strings, replace_strings

def extract_from_external_files(game_dir):
    import yaml
    print("\n  >>> 扫描外部数据文件 (.json/.yml)...")
    found_strings = []
    file_count = 0
    def recursive_find(data):
        if isinstance(data, str): found_strings.append(data)
        elif isinstance(data, list):
            for item in data: recursive_find(item)
        elif isinstance(data, dict):
            for value in data.values(): recursive_find(value)
    for root, dirs, files in os.walk(game_dir):
        if 'tl' in os.path.relpath(root, game_dir).split(os.sep): continue
        for file in files:
            ext = os.path.splitext(file)[1].lower()
            try:
                data = None
                if ext == '.json':
                    with open(os.path.join(root, file), 'r', encoding='utf-8') as f: data = json.load(f)
                elif ext in ['.yaml', '.yml']:
                    with open(os.path.join(root, file), 'r', encoding='utf-8') as f: data = yaml.safe_load(f)
                if data: 
                    file_count += 1
                    recursive_find(data)
            except: pass
    return found_strings, file_count

def extract_deep_python_strings(file_path):
    found = []
    try:
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f: content = f.read()
        matches = re.findall(r'(["\'])((?:\\\1|.)*?)\1', content)
        for _, s in matches: found.append(unescape_string(s))
    except: pass
    return found

def extract_existing_translations(tl_dir):
    existing_set = set()
    if not os.path.exists(tl_dir): return existing_set
    for root, _, files in os.walk(tl_dir):
        for file in files:
            if file.endswith('.rpy'):
                try:
                    with open(os.path.join(root, file), 'r', encoding='utf-8', errors='replace') as f:
                        olds = re.findall(r'^\s*old\s+(["\'])((?:\\\1|.)*?)\1', f.read(), re.MULTILINE)
                        for _, s in olds: existing_set.add(unescape_string(s))
                except: pass
    return existing_set

def save_to_excel_smart(strings, path, headers, counter_obj=None, description_tag=None):
    if not strings: return
    data_rows = []
    for s in strings:
        count = counter_obj[s] if counter_obj else 1
        row = [s, ''] 
        if description_tag:
            row.append(description_tag)
            row.append(count)
            cols = [headers[0], headers[1], headers[2], 'Frequency']
        else:
            row.append(count)
            cols = [headers[0], headers[1], 'Frequency']
        data_rows.append(row)
    data_rows.sort(key=lambda x: x[-1], reverse=True)
    df = pd.DataFrame(data_rows, columns=cols)
    df.to_excel(path, index=False)
    try:
        wb = load_workbook(path); ws = wb.active
        ws.column_dimensions['A'].width = 40; ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 15
        wb.save(path)
    except: pass
    print(f"✔ 保存 Excel: {os.path.basename(path)} (共 {len(strings)} 条)")

def generate_emoji_tables(target_strings, out_dir):
    print(f"\n--- 生成 Tag 保护表 ---")
    patterns = re.compile(r'(\{[^{}]+\}|\[[^\[\]]+\])')
    found_tags = set()
    for s in target_strings:
        matches = patterns.findall(s)
        for m in matches: found_tags.add(m)
    if not found_tags: return
    found_list = sorted(list(found_tags), key=lambda x: len(x), reverse=True)
    wb_pre = Workbook(); ws_pre = wb_pre.active; ws_pre.title = "译前替换(禁翻)"
    wb_post = Workbook(); ws_post = wb_post.active; ws_post.title = "译后还原"
    ws_pre.append(["原文(Tag)", "替换文(Emoji)"]); ws_post.append(["原文(Emoji)", "替换文(Tag)"])
    for i, text in enumerate(found_list):
        emoji = GRS_EMBEDDED_EMOJIS[i % len(GRS_EMBEDDED_EMOJIS)]
        if i >= len(GRS_EMBEDDED_EMOJIS): emoji += GRS_EMBEDDED_EMOJIS[(i // len(GRS_EMBEDDED_EMOJIS)) % len(GRS_EMBEDDED_EMOJIS)]
        ws_pre.append([text, emoji]); ws_post.append([emoji, text])
    if not os.path.exists(out_dir): os.makedirs(out_dir)
    wb_pre.save(os.path.join(out_dir, 'Tag_Protection_Pre(译前).xlsx'))
    wb_post.save(os.path.join(out_dir, 'Tag_Protection_Post(译后).xlsx'))
    print(f"✔ 保护表已生成至: {out_dir}")

def generate_report(counter, output_path):
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=== Ren'Py 文本提取统计报告 ===\n")
        f.write(f"{'次数':<8} | {'文本内容'}\n" + "-" * 60 + "\n")
        for text, count in counter.most_common(300):
            if count > 1: f.write(f"{count:<8} | {text.replace(chr(10), '\\n')}\n")

def generate_rpy_file(strings, output_path, lang_folder):
    if not strings: return
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(f"translate {lang_folder} strings:\n\n")
        for s in strings:
            escaped = s.replace('"', '\\"')
            f.write(f'    old "{escaped}"\n    new ""\n\n')
    print(f"✔ 生成 RPY: {os.path.basename(output_path)}")

def generate_replace_rpy(strings, output_path, lang_folder):
    if not strings: return
    sorted_strings = sorted(list(set(strings)), key=len, reverse=True)
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("init python:\n")
        f.write(f'    if preferences.language == "{lang_folder}":\n')
        f.write("        def replace_text(s):\n")
        f.write("            if not isinstance(s, str): return s\n")
        for s in sorted_strings:
            escaped = s.replace('\\', '\\\\').replace('"', '\\"')
            f.write(f'            s = s.replace("{escaped}", "{escaped}")\n') 
        f.write("            return s\n")
        f.write("        config.replace_text = replace_text\n")

def core_logic_extract(game_dir, lang, mode_choice, need_emoji, need_ai):
    import yaml
    
    tl_lang_dir = os.path.join(game_dir, 'tl', lang)
    base_dir = os.getcwd()
    out_base = os.path.join(base_dir, 'translate_output')
    out_excel = os.path.join(out_base, '1_Excels')
    out_rpy = os.path.join(out_base, '2_RPY_Files')
    out_tools = os.path.join(out_base, '3_Emoji_Tools')
    
    for p in [out_excel, out_rpy, out_tools]:
        if not os.path.exists(p): os.makedirs(p)

    print("\n[步骤 1] 扫描与提取...")
    rpy_files = []
    for root, _, files in os.walk(game_dir):
        if 'tl' in os.path.relpath(root, game_dir).split(os.sep): continue
        for f in files:
            if f.endswith('.rpy'): rpy_files.append(os.path.join(root, f))
    
    raw_names, raw_others, raw_replace = [], [], []
    for f in rpy_files:
        n, t, v, r = extract_strings_from_rpy(f)
        raw_names.extend(n); raw_others.extend(t + v); raw_replace.extend(r)

    ext_file_count = 0
    if mode_choice in ['2', '3']:
        ext_strs, count = extract_from_external_files(game_dir)
        raw_replace.extend(ext_strs)
        ext_file_count = count
    
    if mode_choice == '3':
        print("  >>> 疯狗模式深度扫描...")
        for f in rpy_files: raw_replace.extend(extract_deep_python_strings(f))

    print("\n[步骤 2] 过滤与频率分析...")
    f_names = filter_extracted_strings(raw_names)
    f_others = filter_extracted_strings(raw_others)
    strict = True if mode_choice == '3' else False
    f_replace = filter_extracted_strings(raw_replace, strict_mode=strict)

    all_text_counter = Counter(f_names + f_others + f_replace)
    generate_report(all_text_counter, os.path.join(out_base, 'Duplicate_Report.txt'))

    print("\n[步骤 3] 生成文件...")
    existing_strings = extract_existing_translations(tl_lang_dir)
    
    def finalize_list(raw_list):
        return sorted(list(set([s for s in raw_list if s not in existing_strings])))

    final_names = finalize_list(f_names)
    final_others = finalize_list(f_others)
    final_others = [s for s in final_others if s not in final_names]
    
    standard_pool = set(final_names + final_others)
    final_replace = sorted(list(set([s for s in f_replace if s not in existing_strings and s not in standard_pool])))

    save_to_excel_smart(final_names, os.path.join(out_excel, 'names.xlsx'), 
                        ['Original Text (原文)', 'Translation (在此填入译文)', 'Description (备注)'], 
                        all_text_counter, description_tag='character')
    
    save_to_excel_smart(final_others, os.path.join(out_excel, 'others.xlsx'), 
                        ['Original', 'Translation'], all_text_counter)
    
    save_to_excel_smart(final_replace, os.path.join(out_excel, 'replace_text.xlsx'), 
                        ['Text', 'Replacement'], all_text_counter)
    
    if need_ai and final_names:
        with open(os.path.join(out_base, 'AI_Prompt_Names.txt'), 'w', encoding='utf-8') as f:
            f.write("请翻译以下游戏角色名/专有名词：\n")
            for n in final_names: f.write(f"{n}\n")
        print(f"✔ 生成 AI 提示词: AI_Prompt_Names.txt")

    generate_rpy_file(final_names, os.path.join(out_rpy, 'translate_names.rpy'), lang)
    generate_rpy_file(final_others, os.path.join(out_rpy, 'translate_others.rpy'), lang)
    if final_replace: generate_replace_rpy(final_replace, os.path.join(out_rpy, 'replace.rpy'), lang)

    if need_emoji:
        all_targets = final_names + final_others + final_replace
        generate_emoji_tables(all_targets, out_tools)

    print("\n" + "="*30)
    print("全部完成！")
    
    return {
        'files': len(rpy_files) + ext_file_count,
        'names': len(final_names),
        'others': len(final_others),
        'replace': len(final_replace),
        'total': len(final_names) + len(final_others) + len(final_replace)
    }

if __name__ == "__main__":
    root = tk.Tk()
    app = OldCatApp(root)
    root.mainloop()
